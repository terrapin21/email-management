"""
メール添付書類の解析・情報抽出・Excel書き込みサービス
"""
import io
import json
import logging
import os
import re
import tempfile
from pathlib import Path
from typing import Optional

import anthropic
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app import models
from app.config import settings
from app.services.archive_service import (
    extract_password_from_text,
    find_password_email,
    _extract_zip,
    _extract_7z,
)

logger = logging.getLogger(__name__)

ATTACHMENT_BASE = Path("/app/attachments")
ATT_CACHE = ATTACHMENT_BASE / "cache"


def _resolve_att_path(att) -> str | None:
    """添付ファイルの実ファイルパスを返す。file_pathが未設定ならキャッシュパスを試みる。"""
    if att.file_path and os.path.exists(att.file_path):
        return att.file_path
    cache = ATT_CACHE / f"{att.id}.bin"
    if cache.exists():
        return str(cache)
    return None

MAP_KEYWORDS = ["地図", "案内図", "付近地図", "map", "邸案内図", "周辺図", "ちず", "annai"]
REQUEST_KEYWORDS = ["回収依頼", "養生回収", "養生材回収", "依頼票", "依頼書", "床養生", "リユース"]
_SITE_ID_KEYWORDS = ["コード", "施主", "工事番号", "現場コード", "現場no", "発注no", "案件no", "物件コード",
                     "site id", "site no", "site code", "genba no"]
_SOONEST_KEYWORDS = ["最短", "最速", "最短日", "最短で", "最短納品", "最短回収"]

_ai_client: Optional[anthropic.Anthropic] = None


def _get_ai_client() -> anthropic.Anthropic:
    global _ai_client
    if _ai_client is None:
        _ai_client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
    return _ai_client


def is_map_file(filename: str) -> bool:
    lower = filename.lower()
    return any(kw.lower() in lower for kw in MAP_KEYWORDS)


def is_request_file(filename: str) -> bool:
    return any(kw in filename for kw in REQUEST_KEYWORDS)


def get_file_kind(filename: str) -> str:
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xls")):
        return "excel"
    if lower.endswith(".pdf"):
        return "pdf"
    if lower.endswith((".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".webp")):
        return "image"
    if lower.endswith((".zip", ".7z", ".lzh", ".zi_")):
        return "archive"
    return "other"


# ── テキスト抽出（Excel用） ────────────────────────────────────────────────────

def _parse_excel_text(file_path: str) -> str:
    """Excelをテキストに変換（書類解析と同じロジック）"""
    import datetime as dt
    from openpyxl.utils.datetime import from_excel
    _DATE_RE = re.compile(r'y{1,4}|m{1,5}|d{1,4}|年|月|日', re.IGNORECASE)
    _DATE_MIN, _DATE_MAX = 25569, 73050

    def _cell_val(cell) -> str:
        val = cell.value
        if val is None:
            return ""
        if isinstance(val, (dt.datetime, dt.date)):
            return f"{val.year}　{val.month}　{val.day}"
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            try:
                if cell.is_date:
                    d = from_excel(val)
                    return f"{d.year}　{d.month}　{d.day}"
            except Exception:
                pass
            fmt = cell.number_format or ""
            if _DATE_RE.search(fmt):
                try:
                    d = from_excel(val)
                    return f"{d.year}　{d.month}　{d.day}"
                except Exception:
                    pass
            if _DATE_MIN <= int(val) <= _DATE_MAX:
                try:
                    d = from_excel(val)
                    return f"{d.year}　{d.month}　{d.day}"
                except Exception:
                    pass
        return str(val)

    try:
        with open(file_path, "rb") as f:
            raw = f.read()
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"シート: {ws.title}")
            for row in ws.iter_rows():
                cells = [_cell_val(c) for c in row]
                line = "\t".join(c for c in cells if c.strip())
                if line:
                    lines.append(line)
        return "\n".join(lines).strip()
    except Exception as e:
        logger.warning(f"Excel変換エラー: {e}")
        return ""


# ── AI 抽出 ───────────────────────────────────────────────────────────────────

def _field_label(f: models.ExtractionField) -> str:
    aliases = f.aliases or []
    if aliases:
        return f"{f.field_name}（別の呼び方: {'、'.join(aliases)}）"
    return f.field_name


def _build_extraction_prompt(config: models.MakerExtractionConfig) -> str:
    import datetime as _dt
    today = _dt.date.today()
    current_year = today.year
    next_year = today.year + 1

    required = [f for f in config.fields if f.required]
    optional = [f for f in config.fields if not f.required]
    fields_desc = "必須: " + ", ".join(_field_label(f) for f in required)
    if optional:
        fields_desc += "\n任意: " + ", ".join(_field_label(f) for f in optional)
    all_names = [f.field_name for f in config.fields]
    example = "{" + ", ".join(f'"{n}": "値"' for n in all_names[:3]) + "}"
    return f"""あなたは回収依頼書類から情報を抽出するアシスタントです。
メーカー: {config.maker_name}
抽出項目:
{fields_desc}

書類の内容を解析して、抽出項目の値をJSONのみで返してください。
別の呼び方が記載されていても同じ項目として扱い、JSONのキーは必ず左側の名前を使用してください。
見つからない場合は null を使用。日付は YYYY-M-D 形式（例: 2026-1-5）。ゼロ埋めは不要。
【日付の年について】今日は {today.strftime('%Y年%m月%d日')} です。書類に年の記載がない場合（例: 6月2日）は {current_year}年 を補完してください。ただし {current_year}年 で補完した日付が今日より過去になる場合は {next_year}年 を使用してください。
【重要】日付フィールドに具体的な日付がなく、「最短」「最速」「最短日」「できるだけ早く」などの表現がある場合は、そのフィールドの値を「最短」と設定してください。
例: {example}
JSON以外は絶対に出力しないでください。"""


def _extract_from_body(body_text: str, config: models.MakerExtractionConfig) -> dict:
    """メール本文からフィールドを抽出する"""
    ai = _get_ai_client()
    system = _build_extraction_prompt(config)
    try:
        resp = ai.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=800,
            system=system,
            messages=[{"role": "user", "content": f"以下のメール本文から情報を抽出してください:\n\n{body_text[:5000]}"}],
        )
        return _parse_ai_response(resp.content[0].text)
    except Exception as e:
        logger.warning(f"本文抽出エラー: {e}")
        return {}


def _parse_ai_response(text: str) -> dict:
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except Exception:
            pass
    return {}


def analyze_file(file_path: str, filename: str, config: models.MakerExtractionConfig) -> dict:
    """書類解析システムと同じAPI呼び出しでファイルを解析し、設定フィールドを抽出する"""
    import base64
    ext = Path(filename).suffix.lower()
    ai = _get_ai_client()
    system = _build_extraction_prompt(config)
    user_text = "この書類から指定された項目をJSONで抽出してください。"

    try:
        if ext in (".xlsx", ".xls"):
            text = _parse_excel_text(file_path)
            if not text:
                return {}
            resp = ai.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=800,
                system=system,
                messages=[{"role": "user", "content": f"以下のExcel内容から情報を抽出してください:\n\n{text[:5000]}"}],
            )
            return _parse_ai_response(resp.content[0].text)

        with open(file_path, "rb") as f:
            b64 = base64.standard_b64encode(f.read()).decode()

        if ext == ".pdf":
            resp = ai.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=800,
                system=system,
                messages=[{"role": "user", "content": [
                    {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                    {"type": "text", "text": user_text},
                ]}],
                extra_headers={"anthropic-beta": "pdfs-2024-09-25"},
            )
            return _parse_ai_response(resp.content[0].text)

        media_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
                     ".gif": "image/gif", ".bmp": "image/bmp", ".webp": "image/webp",
                     ".tiff": "image/jpeg"}
        mime = media_map.get(ext)
        if mime:
            resp = ai.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=800,
                system=system,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
                    {"type": "text", "text": user_text},
                ]}],
            )
            return _parse_ai_response(resp.content[0].text)

    except Exception as e:
        logger.warning(f"書類解析エラー ({filename}): {e}")
    return {}


def analyze_file_with_confidence(file_path: str, filename: str, config: models.MakerExtractionConfig) -> tuple[dict, bool]:
    """画像ファイルを3回解析して信頼度を確認する"""
    ext = Path(filename).suffix.lower()
    media_map = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tiff"}
    if ext not in media_map:
        result = analyze_file(file_path, filename, config)
        required = [f.field_name for f in config.fields if f.required]
        confident = bool(result) and all(result.get(f) for f in required)
        return result, confident

    results = [analyze_file(file_path, filename, config) for _ in range(3)]
    if not any(results):
        return {}, False

    first = results[0]
    if all(r == first for r in results[1:]) and first:
        return first, True

    merged = {}
    for key in first:
        vals = [r.get(key) for r in results]
        if len(set(str(v) for v in vals)) == 1:
            merged[key] = first[key]

    required = [f.field_name for f in config.fields if f.required]
    confident = bool(merged) and all(merged.get(f) for f in required)
    return merged, confident


# ── ローカルファイル操作（NASはWindowsからrobocopyで同期） ────────────────────

LOCAL_OUTPUT_BASE = Path("/app/nas_output")


def _normalize_date_value(value: str) -> tuple[str, bool]:
    """
    抽出された日付文字列を正規化する。
    Returns: (正規化後の文字列, 今日以降かどうか)
    - 年なし（M月D日, M/D 等）→ 今年で試し、過去なら来年
    - 年あり → そのまま検証（過去ならFalse）
    """
    import datetime as _dt
    today = _dt.date.today()

    s = str(value).strip()
    month = day = year = None

    def _try_date(y: int, mo: int, d: int) -> "_dt.date | None":
        try:
            return _dt.date(y, mo, d)
        except Exception:
            return None

    # YYYY-MM-DD / YYYY/MM/DD
    m = re.match(r'^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$', s)
    if m:
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))

    if year is None:
        # YYYY年M月D日
        m = re.match(r'^(\d{4})年(\d{1,2})月(\d{1,2})日', s)
        if m:
            year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))

    if year is None:
        # M月D日（年なし）
        m = re.match(r'^(\d{1,2})月(\d{1,2})日', s)
        if m:
            month, day = int(m.group(1)), int(m.group(2))

    if month is None:
        # M/D or M-D（年なし）
        m = re.match(r'^(\d{1,2})[/\-](\d{1,2})$', s)
        if m:
            month, day = int(m.group(1)), int(m.group(2))

    if month is None:
        # 解析不能：そのまま返す（バリデーションしない）
        return s, True

    if year is not None:
        # 年あり: そのまま検証
        d = _try_date(year, month, day)
        if d:
            return f"{d.year}-{d.month}-{d.day}", d >= today
        return s, False
    else:
        # 年なし: 今年→来年の順で試す
        for y in (today.year, today.year + 1):
            d = _try_date(y, month, day)
            if d and d >= today:
                return f"{d.year}-{d.month}-{d.day}", True
        return s, False


def _has_soonest_keyword(text: str) -> bool:
    return any(kw in (text or "") for kw in _SOONEST_KEYWORDS)


def _find_empty_date_field(extracted_data: dict, config: models.MakerExtractionConfig) -> str | None:
    """値が空または「最短」の日付フィールド名を返す"""
    for field in sorted(config.fields, key=lambda f: f.order):
        if field.field_type == "date":
            val = extracted_data.get(field.field_name)
            if not val or _has_soonest_keyword(str(val)):
                return field.field_name
    return None


def _is_site_id_field(field_name: str) -> bool:
    n = field_name.lower()
    return any(k in n for k in _SITE_ID_KEYWORDS)


def _collect_map_files(
    email: models.Email,
    extracted_data: dict,
    config: models.MakerExtractionConfig,
    db: Session,
) -> list[tuple[str, str]]:
    """メール本体・関連メール双方から地図ファイルを収集して返す"""
    map_files: list[tuple[str, str]] = []
    for att in (email.attachments or []):
        if is_map_file(att.filename):
            fpath = _resolve_att_path(att)
            if fpath:
                map_files.append((fpath, att.filename))
    if not map_files:
        map_files = _find_map_from_related_emails(email, extracted_data, config, db)
    return map_files


def _find_map_from_related_emails(
    email: models.Email,
    extracted_data: dict,
    config: models.MakerExtractionConfig,
    db: Session,
) -> list[tuple[str, str]]:
    """同じコード値を持つ関連メールの添付ファイルから地図ファイルを探す"""
    # 抽出データのコード系フィールド値を収集
    code_values = []
    for field in config.fields:
        if field.field_type == "code":
            val = extracted_data.get(field.field_name)
            if val:
                code_values.append(str(val))

    # コード系フィールドが設定されていない場合はEmailFieldからも探す
    if not code_values:
        for ef in (email.extracted_fields or []):
            if _is_site_id_field(ef.field_name) and ef.field_value:
                code_values.append(ef.field_value)

    if not code_values:
        return []

    # 同じコード値を持つ他のメールを取得
    related_fields = db.query(models.EmailField).filter(
        models.EmailField.field_value.in_(code_values),
        models.EmailField.email_id != email.id,
    ).all()

    if not related_fields:
        return []

    related_email_ids = list({rf.email_id for rf in related_fields})

    # 関連メールの添付ファイルから地図ファイル（画像）を探す
    IMAGE_EXTS = {".png", ".tif", ".tiff", ".jpg", ".jpeg", ".bmp", ".webp"}
    map_files = []
    for rel_email_id in related_email_ids:
        rel_email = db.query(models.Email).get(rel_email_id)
        if not rel_email:
            continue
        for att in (rel_email.attachments or []):
            ext = Path(att.filename).suffix.lower()
            if is_map_file(att.filename) or ext in IMAGE_EXTS:
                fpath = _resolve_att_path(att)
                if fpath:
                    map_files.append((fpath, att.filename))
                    logger.info(f"関連メール(id={rel_email_id})から地図ファイル取得: {att.filename}")

    return map_files


def _nas_to_local(nas_path: str) -> Path:
    """NASパス(UNCパス)をDockerローカルパスに変換する"""
    # バックスラッシュをスラッシュに統一
    normalized = nas_path.replace("\\", "/").lstrip("/")
    # //host/share/... の場合、host/share部分を除去してベースに結合
    parts = normalized.split("/")
    if len(parts) >= 2 and "." in parts[0]:
        # //192.168.x.x/share/path → skip host + share
        rel = "/".join(parts[2:])
    else:
        rel = normalized
    return LOCAL_OUTPUT_BASE / rel


def _safe_filename(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|\s]', "_", str(value or "unknown"))


def write_excel_row(data: dict, config: models.MakerExtractionConfig) -> bool:
    if not config.excel_file_path:
        logger.warning("Excel パスが未設定")
        return False

    sorted_fields = sorted(config.fields, key=lambda f: f.order)
    headers = [f.field_name for f in sorted_fields] + ["処理済"]
    row_values = [str(data.get(f.field_name) or "") for f in sorted_fields] + [""]

    try:
        local_path = _nas_to_local(config.excel_file_path)
        local_path.parent.mkdir(parents=True, exist_ok=True)

        if local_path.exists():
            wb = load_workbook(io.BytesIO(local_path.read_bytes()))
            ws = wb.active
            # 既存ファイルに「処理済」列がなければ追加
            existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            if "処理済" not in existing_headers:
                col = ws.max_column + 1
                ws.cell(1, col).value = "処理済"
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, col).value = ""
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)

        ws.append(row_values)
        buf = io.BytesIO()
        wb.save(buf)
        local_path.write_bytes(buf.getvalue())

        logger.info(f"Excel 書き込み完了: {local_path}")

        # エクセルと同じディレクトリにトリガーファイルを書く（PAD連携用）
        # robocopyがエクセルと一緒にNASへ同期する
        _write_pad_trigger(local_path.parent)

        return True
    except Exception as e:
        logger.error(f"Excel 書き込みエラー: {e}")
        return False


def _write_pad_trigger(directory: Path) -> None:
    """PAD連携用トリガーファイルをエクセルと同じディレクトリに書く"""
    import datetime
    trigger_path = directory / "_pad_trigger.txt"
    try:
        trigger_path.write_text(
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            encoding="utf-8",
        )
        logger.info(f"PADトリガー書き込み: {trigger_path}")
    except Exception as e:
        logger.warning(f"PADトリガー書き込み失敗: {e}")


def save_maps_to_nas(map_file_paths: list[tuple[str, str]], data: dict, config: models.MakerExtractionConfig) -> bool:
    """複数の地図画像を1つのPDFにまとめてNASに保存する（1枚でもPDF化）"""
    if not config.map_save_path:
        logger.warning("地図保存パスが未設定")
        return False
    if not map_file_paths:
        return False

    from PIL import Image as PilImage

    date_field = config.map_date_field or "回収日"
    code_val = None
    for f in config.fields:
        if f.field_type == "code" or f.field_name in ("コード",):
            code_val = data.get(f.field_name)
            if code_val:
                break
    code = code_val or data.get("コード") or "unknown"
    date_val = data.get(date_field) or "unknown"
    save_name = f"{_safe_filename(code)}_{_safe_filename(date_val)}_案内図.pdf"

    IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tiff", ".tif", ".webp"}
    images = []
    for fpath, fname in map_file_paths:
        ext = Path(fname).suffix.lower()
        if ext not in IMAGE_EXTS:
            logger.info(f"地図ファイルをスキップ（非画像）: {fname}")
            continue
        try:
            img = PilImage.open(fpath).convert("RGB")
            images.append(img)
            logger.info(f"地図画像を追加: {fname}")
        except Exception as e:
            logger.warning(f"地図画像読み込みエラー ({fname}): {e}")

    if not images:
        logger.warning("有効な地図画像が見つかりませんでした")
        return False

    try:
        local_dir = _nas_to_local(config.map_save_path)
        local_dir.mkdir(parents=True, exist_ok=True)
        dest = local_dir / save_name
        images[0].save(str(dest), "PDF", resolution=150, save_all=True, append_images=images[1:])
        logger.info(f"地図PDF保存完了: {dest} ({len(images)}ページ)")
        return True
    except Exception as e:
        logger.error(f"地図PDF保存エラー: {e}")
        return False


def save_map_to_nas(file_path: str, filename: str, data: dict, config: models.MakerExtractionConfig) -> bool:
    """後方互換用。単一ファイルをsave_maps_to_nasに委譲する"""
    return save_maps_to_nas([(file_path, filename)], data, config)


# ── メイン処理 ────────────────────────────────────────────────────────────────

def process_email_extraction(email_id: int, db: Session) -> dict:
    email = db.query(models.Email).filter(models.Email.id == email_id).first()
    if not email:
        logger.warning(f"抽出スキップ: email_id={email_id} メールが見つかりません")
        return {"success": False, "reason": "メールが見つかりません"}

    maker = email.ai_manufacturer
    logger.info(f"抽出開始: email_id={email_id}, ai_manufacturer={maker!r}")
    if not maker:
        logger.warning(f"抽出スキップ: email_id={email_id} ai_manufacturerが未設定")
        return {"success": False, "reason": "メーカー情報が未解析です"}

    # 設定名がAI名に含まれる、またはAI名が設定名に含まれる、どちらでも一致とみなす
    maker_lower = maker.lower()
    all_configs = db.query(models.MakerExtractionConfig).all()
    config = next(
        (c for c in all_configs
         if c.maker_name.lower() in maker_lower or maker_lower in c.maker_name.lower()),
        None,
    )
    if not config:
        logger.warning(f"抽出スキップ: email_id={email_id} メーカー「{maker}」の設定なし")
        return {"success": False, "reason": f"メーカー「{maker}」の抽出設定がありません"}

    # 既存結果を削除して再処理
    db.query(models.ExtractionResult).filter(
        models.ExtractionResult.email_id == email_id
    ).delete()

    attachments = email.attachments or []
    extracted_data: dict = {}
    attachment_pattern = "unknown"
    status = "needs_review"
    review_reason = ""
    map_file_paths: list[tuple[str, str]] = []  # (file_path, filename)

    if attachments:
        map_atts = [a for a in attachments if is_map_file(a.filename)]
        req_atts = [a for a in attachments if not is_map_file(a.filename)]

        processed_any = False
        for req_att in req_atts:
            fpath = _resolve_att_path(req_att)
            if not fpath:
                logger.warning(f"添付ファイルが見つかりません: id={req_att.id} filename={req_att.filename}")
                continue

            kind = get_file_kind(req_att.filename)

            if kind in ("excel", "pdf", "image"):
                attachment_pattern = f"case{'1' if kind != 'image' else '2'}"
                processed_any = True
                data, confident = analyze_file_with_confidence(fpath, req_att.filename, config)
                if data:
                    extracted_data = data
                    if confident:
                        status = "completed"
                    else:
                        status = "needs_review"
                        review_reason = "解析結果の信頼度が不足しています（必須項目が取得できませんでした）"
                else:
                    review_reason = "AIが書類から情報を抽出できませんでした"

            elif kind == "archive":
                attachment_pattern = "case3"
                processed_any = True

                # 解凍済みアーカイブレコードがあればそのファイルを直接使う
                arch_record = next(
                    (a for a in (email.encrypted_archives or [])
                     if a.attachment_id == req_att.id
                     and a.status == models.ArchiveStatusEnum.extracted),
                    None,
                )
                if arch_record and arch_record.extracted_files:
                    logger.info(f"解凍済みアーカイブから抽出: archive_id={arch_record.id}, "
                                f"files={[ef.filename for ef in arch_record.extracted_files]}")
                    for ef in arch_record.extracted_files:
                        if not os.path.exists(ef.file_path):
                            logger.warning(f"解凍済みファイルが見つかりません: {ef.file_path}")
                            continue
                        if is_map_file(ef.filename):
                            map_file_paths.append((ef.file_path, ef.filename))
                        elif get_file_kind(ef.filename) in ("excel", "pdf", "image"):
                            data, confident = analyze_file_with_confidence(ef.file_path, ef.filename, config)
                            if data:
                                extracted_data = data
                                status = "completed" if confident else "needs_review"
                                if not confident:
                                    review_reason = "解析結果の信頼度が不足しています（必須項目が取得できませんでした）"
                            else:
                                review_reason = "AIが解凍済みファイルから情報を抽出できませんでした"
                else:
                    # 未解凍: パスワードメールを探して解凍を試みる
                    pw_email = find_password_email(db, email)
                    password = extract_password_from_text(pw_email.body_text or "") if pw_email else None

                    if password:
                        try:
                            with open(fpath, "rb") as f:
                                archive_data = f.read()
                            with tempfile.TemporaryDirectory() as tmpdir:
                                tmp_path = Path(tmpdir)
                                lower = req_att.filename.lower()
                                if lower.endswith(".zip"):
                                    files = _extract_zip(archive_data, password, tmp_path)
                                elif lower.endswith(".7z"):
                                    files = _extract_7z(archive_data, password, tmp_path)
                                else:
                                    files = []

                                for file_info in files:
                                    fname = file_info["filename"]
                                    fpath_inner = file_info["file_path"]
                                    if is_map_file(fname):
                                        import shutil
                                        dest = ATTACHMENT_BASE / "extracted" / f"map_{email_id}_{fname}"
                                        dest.parent.mkdir(parents=True, exist_ok=True)
                                        shutil.copy2(fpath_inner, dest)
                                        map_file_paths.append((str(dest), fname))
                                    elif get_file_kind(fname) in ("excel", "pdf", "image"):
                                        data, confident = analyze_file_with_confidence(fpath_inner, fname, config)
                                        if data and not extracted_data:
                                            extracted_data = data
                                            status = "completed" if confident else "needs_review"
                                            if not confident:
                                                review_reason = "解析結果の信頼度が不足しています"
                        except Exception as e:
                            review_reason = f"圧縮ファイルの解凍に失敗しました: {e}"
                    else:
                        review_reason = "パスワードメールが見つかりません（PPAP対応待ち）"

            else:
                attachment_pattern = "case4"
                review_reason = "DLリンク形式のため手動対応が必要です"

            if status == "completed":
                break

        if not processed_any:
            review_reason = "依頼書類の添付が見つかりませんでした（ファイルが存在しないか未サポート形式）"

        # 地図ファイルを収集（本体添付 → 関連メールの順に探す）
        map_file_paths = _collect_map_files(email, extracted_data, config, db)

        # 添付解析後にメール本文でも抽出して不足フィールドを補完
        if email.body_text:
            body_data = _extract_from_body(email.body_text, config)
            for k, v in body_data.items():
                if v and not extracted_data.get(k):
                    extracted_data[k] = v
            # 補完後に必須フィールドが揃えば completed に昇格
            required_names = [f.field_name for f in config.fields if f.required]
            if status != "completed" and extracted_data and all(extracted_data.get(n) for n in required_names):
                status = "completed"
                review_reason = ""

    else:
        # 添付なし: メール本文のみから抽出
        attachment_pattern = "pattern2"
        if email.body_text:
            extracted_data = _extract_from_body(email.body_text, config)
            required_names = [f.field_name for f in config.fields if f.required]
            if extracted_data and all(extracted_data.get(n) for n in required_names):
                status = "completed"
            elif extracted_data:
                status = "needs_review"
                review_reason = "必須項目の一部が取得できませんでした"
            else:
                review_reason = "本文から情報を抽出できませんでした"
        else:
            review_reason = "メール本文が空です"

        # 地図ファイルを収集（本体添付 → 関連メールの順に探す）
        if extracted_data:
            map_file_paths = _collect_map_files(email, extracted_data, config, db)

    # 地図必須チェック
    if config.map_required and not map_file_paths and status == "completed":
        status = "needs_review"
        review_reason = "地図ファイルが見つかりませんでした（必須設定）"

    # 日付フィールドの正規化・過去日付チェック
    for field in config.fields:
        if field.field_type != "date":
            continue
        val = extracted_data.get(field.field_name)
        if not val or _has_soonest_keyword(str(val)):
            continue
        normalized, is_future = _normalize_date_value(str(val))
        extracted_data[field.field_name] = normalized
        if not is_future:
            status = "needs_review"
            review_reason = f"{field.field_name} が過去の日付（{normalized}）のため要確認です"

    # 最短キーワードチェック：本文・抽出値・日付フィールドに「最短」があれば手動入力待ち
    needs_soonest_date = False
    soonest_date_field = None
    body_text = email.body_text or ""
    extracted_values_text = " ".join(str(v) for v in extracted_data.values() if v)
    has_soonest = (
        _has_soonest_keyword(body_text)
        or _has_soonest_keyword(extracted_values_text)
    )
    if has_soonest:
        soonest_date_field = _find_empty_date_field(extracted_data, config)
        if soonest_date_field:
            needs_soonest_date = True
            status = "needs_review"
            # 「最短」が日付フィールドの値として入っている場合はクリアしてから手動入力待ちにする
            if _has_soonest_keyword(str(extracted_data.get(soonest_date_field) or "")):
                extracted_data[soonest_date_field] = None
            review_reason = f"「最短」指定のため {soonest_date_field} を手動入力してください"

    # Excel書き込みと地図保存（最短入力待ちの場合は書き込みをスキップ）
    excel_written = False
    if status == "completed" and extracted_data and not needs_soonest_date:
        excel_written = write_excel_row(extracted_data, config)
        if not excel_written:
            status = "needs_review"
            review_reason = "Excel への書き込みに失敗しました（NAS接続を確認してください）"

        if excel_written and map_file_paths:
            save_maps_to_nas(map_file_paths, extracted_data, config)

    # 結果を保存
    result = models.ExtractionResult(
        email_id=email_id,
        config_id=config.id,
        extracted_data=extracted_data,
        status=status,
        review_reason=review_reason or None,
        attachment_pattern=attachment_pattern,
        excel_written=excel_written,
        needs_soonest_date=needs_soonest_date,
        soonest_date_field=soonest_date_field,
    )
    db.add(result)

    # 要確認ならメールステータスを更新
    if status == "needs_review":
        sr = email.status_record
        if sr:
            sr.status = models.EmailStatusEnum.needs_review
        else:
            db.add(models.EmailStatusRecord(
                email_id=email_id,
                status=models.EmailStatusEnum.needs_review,
            ))

    db.commit()
    logger.info(f"抽出完了: email_id={email_id}, status={status}, pattern={attachment_pattern}")
    return {
        "success": status == "completed",
        "status": status,
        "extracted_data": extracted_data,
        "reason": review_reason,
        "pattern": attachment_pattern,
    }

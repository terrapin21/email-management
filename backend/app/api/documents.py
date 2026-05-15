import base64
import io
import json
import logging
import re
import datetime
import unicodedata
import urllib.parse
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
import anthropic
import openpyxl
from openpyxl.utils.datetime import from_excel
from app.config import settings

# 日付書式を示す文字列パターン（英語・日本語両対応）
_DATE_FMT_RE = re.compile(r'y{1,4}|m{1,5}|d{1,4}|年|月|日', re.IGNORECASE)
# Excelの日付シリアル値の範囲（1970年〜2100年）
_EXCEL_DATE_MIN = 25569
_EXCEL_DATE_MAX = 73050

router = APIRouter(prefix="/api/documents", tags=["documents"])
logger = logging.getLogger(__name__)
client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)

SYSTEM_PROMPT = "あなたはデータ抽出の専門家です。指示に従って情報を正確に抽出し、必ずJSON形式のみで回答してください。余分なテキストや説明は一切含めないでください。"

_ID_RULE = "施主No・施主コード・工事番号・現場コード・発注No・案件No・物件コードなどメーカーによって呼び名が異なる現場固有IDは、値にIDと名称（〇〇様邸・〇〇邸・〇〇様など）が続けて書かれていてもID部分（英数字・ハイフン等で構成されるコード）のみを抽出してください（例：「AABBCC〇〇様邸」→「AABBCC」、「A-1234田中様邸」→「A-1234」）。"

TEXT_PROMPT = lambda text: f"""以下のテキストデータから次の情報を抽出してください。
見つからない場合は空文字で回答してください。
回収日は年・月・日を別々のフィールドに分けて数字のみで抽出してください（例：年=2026、月=4、日=24）。
{_ID_RULE}

テキスト：
---
{text}
---

必ず以下のJSON形式のみで回答してください（他のテキストは絶対に含めないこと）：
{{"施主No": "施主コード/施主番号/工事番号/発注No等のID部分のみ", "回収日_年": "", "回収日_月": "", "回収日_日": ""}}"""

EXCEL_PROMPT = lambda text: f"""以下はExcelの回収依頼票データです。各項目の値を抽出してください。
見つからない場合は空文字にしてください。ふりがなは含めないでください。
回収希望日は年・月・日を別々のフィールドに分けて数字のみで抽出してください（例：年=2026、月=4、日=24）。
{_ID_RULE}

テキスト：
---
{text}
---

必ず以下のJSON形式のみで回答してください（他のテキストは絶対に含めないこと）：
{{"支社名":"","現場名":"","施主コード":"","建築地":"","監督名":"","携帯電話":"","回収希望日_年":"","回収希望日_月":"","回収希望日_日":"","保管場所":"","ゲート鍵番号":"","工事用キー保管場所":""}}"""

def FILE_PROMPT(filename: str) -> str:
    filename_hint = f"\nファイル名：「{filename}」\nファイル名も種類判定の参考にしてください（例：ファイル名に「案内図」とあれば案内図、「図面」とあれば図面の可能性が高い）。\n"
    return f"""このファイルを分析してください。
{filename_hint}
【ステップ1】このファイルの種類を以下から1つ選んで判定してください：
- 案内図：地図・住所・現地への道順・周辺地図など、場所を示すもの。住所や地名、矢印による経路が含まれていれば案内図。
- 図面：平面図。建物の間取り・部屋の配置・寸法などを示す設計図。地図ではなく建物内部の構造を表すもの。
- 回収依頼書：養生材・資材などの回収を依頼する書類。施主コード・回収日・保管場所などの項目が含まれるもの。
- その他：上記のいずれにも当てはまらないもの。

【ステップ2】「回収依頼書」と判定した場合のみ、以下の情報をすべて抽出してください：
- 施主No（施主番号・施主コード・工事番号・現場コード・発注No・案件No・物件コードなどメーカーによって呼び名が異なる現場固有IDを含む）
- 回収日（回収予定日・回収希望日・最短回収なども含む）
- 現場名・建築地・支社名・監督名・担当者名・携帯電話
- 保管場所（養生材・資材の保管場所）
- ゲート鍵番号（ゲートキー番号・鍵番号・キー番号など）
- 工事用キー保管場所

「回収依頼書」以外の場合、施主No・回収日以外のフィールドはすべて空文字としてください。

回収日は年・月・日を別々のフィールドに分けて数字のみで抽出してください（例：年=2026、月=4、日=24）。
{_ID_RULE}

必ず以下のJSON形式のみで回答してください（他のテキストは絶対に含めないこと）：
{{"種類": "案内図|図面|回収依頼書|その他のいずれか", "施主No": "ID部分のみ", "回収日_年": "", "回収日_月": "", "回収日_日": "", "現場名": "", "建築地": "", "支社名": "", "監督名": "", "携帯電話": "", "保管場所": "", "ゲート鍵番号": "", "工事用キー保管場所": ""}}"""


def _to_date_str(val) -> str | None:
    try:
        dt = from_excel(val)
        return f"{dt.year}　{dt.month}　{dt.day}"
    except Exception:
        return None


def _cell_value(cell) -> str:
    val = cell.value
    if val is None:
        return ""
    # openpyxlがdatetimeで返した場合
    if isinstance(val, (datetime.datetime, datetime.date)):
        return f"{val.year}　{val.month}　{val.day}"
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        int_val = int(val)
        # 1. is_dateフラグで判定
        try:
            if cell.is_date:
                result = _to_date_str(val)
                if result:
                    return result
        except Exception:
            pass
        # 2. number_formatの文字列パターンで判定
        fmt = cell.number_format or ""
        if _DATE_FMT_RE.search(fmt):
            result = _to_date_str(val)
            if result:
                return result
        # 3. 値が日付シリアル値の範囲内なら変換を試みる
        if _EXCEL_DATE_MIN <= int_val <= _EXCEL_DATE_MAX:
            result = _to_date_str(val)
            if result:
                return result
    return str(val)


def parse_excel(data: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"シート: {sheet.title}")
        for row in sheet.iter_rows():
            cells = [_cell_value(c) for c in row]
            line = "\t".join(c for c in cells if c.strip())
            if line:
                lines.append(line)
        lines.append("")
    return "\n".join(lines).strip()


def extract_json(text: str) -> dict:
    match = __import__("re").search(r"\{[\s\S]*\}", text)
    if not match:
        raise ValueError("JSONが見つかりません")
    return json.loads(match.group())


_CODE_FIELDS = ("施主No", "施主コード")

def normalize_codes(data: dict) -> dict:
    """施主コード系フィールドの全角英数字を半角に変換する。"""
    for key in _CODE_FIELDS:
        if key in data and data[key]:
            data[key] = unicodedata.normalize("NFKC", data[key])
    return data


MANAGER_PHONES: dict[str, str] = {
    "室賀": "080-6908-7853", "室賀光明": "080-6908-7853",
    "芦田": "090-8446-7808",
    "杉山": "070-1608-6731", "杉山正城": "070-1608-6731",
    "岩瀬": "080-6948-8656", "岩瀬新也": "080-6948-8656",
    "入谷": "090-3565-0887", "入谷太貴": "090-3565-0887",
    "奥": "080-6914-2071", "奥亮二": "080-6914-2071",
    "大家": "090-7439-8819", "大家憲二": "090-7439-8819",
    "尾崎": "090-4212-0018", "尾崎健心": "090-4212-0018",
    "天野": "070-2322-8328", "天野海紀": "070-2322-8328",
    "伊藤": "090-2937-6331",
    "大塚": "090-3869-1952",
    "上浦": "090-1239-4531", "上浦健一": "090-1239-4531",
    "勝田": "080-5155-2454", "勝田竜司": "080-5155-2454",
    "河竹": "080-6979-3105", "河竹敏宏": "080-6979-3105",
    "金子": "080-6948-8494", "金子直史": "080-6948-8494",
    "加藤": "080-5627-4435", "加藤秀": "080-5627-4435",
    "桐畑": "080-6052-2296",
    "釘本": "090-6380-8098",
    "西郷": "090-8188-1864", "西郷眞生": "090-8188-1864",
    "酒井": "090-4233-4529", "酒井拓夢": "090-4233-4529",
    "鈴木": "090-3554-3503", "鈴木伸啓": "090-3554-3503",
    "榊原": "090-1832-4569", "榊原数人": "090-1832-4569",
    "竹田": "080-6972-9486", "竹田光宏": "080-6972-9486",
    "中野": "090-4150-8115", "中野修": "090-4150-8115",
    "内藤": "080-5291-3105", "内藤龍馬": "080-5291-3105",
    "原崎": "090-6334-1355", "原崎大地": "090-6334-1355",
    "兵藤": "090-3553-7398", "兵藤剛友": "090-3553-7398",
    "橋場": "090-1833-3805", "橋場大介": "090-1833-3805",
    "平野": "070-3789-8956", "平野颯也": "070-3789-8956",
    "三田": "090-2943-8528",
    "三鍋": "080-6955-3197", "三鍋一輝": "080-6955-3197",
    "宮本": "070-2232-1675", "宮本暁人": "070-2232-1675",
    "村林": "080-5105-4961", "村林宏紀": "080-5105-4961",
    "吉田": "090-9609-7688",
    "山本": "090-1835-5800", "山本（メンダイ）": "090-1835-5800",
    "山口": "090-8173-8837",
    "渡邉": "090-4212-0250", "渡邉一馬": "090-4212-0250",
    "渡辺": "090-3552-9119", "渡辺究": "090-3552-9119",
}


def _fmt_schedule_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return f"{val.month}月{val.day}日"
    return str(val).strip()


def _is_schedule_data_row(row: tuple) -> bool:
    if len(row) <= 3:
        return False
    order_no = row[2]
    施主名 = row[3]
    if not isinstance(order_no, str) or not order_no.strip():
        return False
    clean = order_no.replace("\n", "").replace("\r", "").strip()
    if "ｵｰﾀﾞｰ" in clean or clean.startswith("<") or clean.startswith("■"):
        return False
    if not isinstance(施主名, str) or not 施主名.strip():
        return False
    return True


def parse_schedule_excel(data: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    records = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if not _is_schedule_data_row(row):
                continue
            order_no = str(row[2]).strip()
            施主名 = str(row[3]).strip()
            管理者_raw = row[7]
            管理者 = str(管理者_raw).strip() if 管理者_raw is not None else ""
            最短 = row[20] if len(row) > 20 else None
            最終 = row[22] if len(row) > 22 else None
            キーBOX = row[23] if len(row) > 23 else None
            phone = MANAGER_PHONES.get(管理者, "")
            records.append({
                "オーダーNo": order_no,
                "施主名": 施主名,
                "管理者": 管理者,
                "管理者電話番号": phone,
                "床養生回収可能日_最短": _fmt_schedule_date(最短),
                "床養生回収可能日_最終": _fmt_schedule_date(最終),
                "キーBOX No": str(キーBOX).strip() if キーBOX is not None else "",
            })
    return records


@router.post("/analyze-schedule")
async def analyze_schedule(file: UploadFile = File(...)):
    try:
        raw = await file.read()
        records = parse_schedule_excel(raw)
        if not records:
            raise HTTPException(status_code=400, detail="有効なデータが見つかりませんでした")

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "社内検査リスト"

        headers = [
            "オーダーNo", "施主名", "管理者", "管理者電話番号",
            "床養生回収可能日（最短）", "床養生回収可能日（最終）", "キーBOX No",
        ]
        ws_out.append(headers)
        for r in records:
            ws_out.append([
                r["オーダーNo"], r["施主名"], r["管理者"], r["管理者電話番号"],
                r["床養生回収可能日_最短"], r["床養生回収可能日_最終"], r["キーBOX No"],
            ])

        # 列幅の調整
        col_widths = [14, 18, 8, 16, 20, 20, 12]
        for i, w in enumerate(col_widths, 1):
            ws_out.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb_out.save(buf)
        buf.seek(0)

        encoded = urllib.parse.quote("社内検査リスト.xlsx")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"社内検査シート解析エラー: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/analyze")
async def analyze_document(
    file: UploadFile = File(None),
    text: str = Form(None),
):
    if not file and not text:
        raise HTTPException(status_code=400, detail="ファイルまたはテキストを指定してください")

    try:
        if text:
            prompt = TEXT_PROMPT(text)
            message = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=512,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": prompt}],
            )
            data = normalize_codes(extract_json(message.content[0].text))
            return {"success": True, "type": "text", "data": data}

        raw = await file.read()
        filename = file.filename or ""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if ext in ("xlsx", "xls"):
            excel_text = parse_excel(raw)
            prompt = EXCEL_PROMPT(excel_text)
            message = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=512,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": prompt}],
            )
            data = normalize_codes(extract_json(message.content[0].text))
            return {"success": True, "type": "excel", "data": data}

        b64 = base64.standard_b64encode(raw).decode()

        file_prompt_text = FILE_PROMPT(filename)

        if ext == "pdf":
            message = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=512,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": [
                    {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                    {"type": "text", "text": file_prompt_text},
                ]}],
                extra_headers={"anthropic-beta": "pdfs-2024-09-25"},
            )
        elif ext in ("jpg", "jpeg", "png"):
            mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
            message = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=512,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
                    {"type": "text", "text": file_prompt_text},
                ]}],
            )
        else:
            raise HTTPException(status_code=400, detail="対応していないファイル形式です")

        data = normalize_codes(extract_json(message.content[0].text))
        return {"success": True, "type": "file", "data": data}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"書類解析エラー: {e}")
        raise HTTPException(status_code=500, detail=str(e))
